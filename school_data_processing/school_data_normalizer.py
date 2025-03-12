import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from fuzzywuzzy import fuzz
from collections import defaultdict
import re
import unicodedata

BASE_DIR = r"C:\Users\admitere.DESKTOP-ECSQHO6\Desktop\Situatii Judete"

# List of cities to remove
CITIES = ["ALBA IULIA", "CÂMPENI", "SEBEȘ", "OCNA MUREȘ", "BLAJ"]


def normalize_text(text):
    """ Normalize the text for comparison (remove diacritics, quotes, special characters). """
    if not text:
        return ""

    text = unicodedata.normalize("NFKD", text)
    text = re.sub(r'[“”„",]', '', text)  # Remove quotes and commas
    text = re.sub(r'\s+', ' ', text).strip()
    text = text.upper()
    return text


def remove_city_names(school_name):
    """ Remove city names from school names for better comparison. """
    for city in CITIES:
        school_name = school_name.replace(city, "").strip()
    return school_name


def is_similar_school(name1, name2):
    """ Check if two schools are similar, including special cases """

    # Remove quotes, apostrophes, and normalize text
    name1_clean = re.sub(r"[\"'’”“„]", '', name1)
    name2_clean = re.sub(r"[\"'’”“„]", '', name2)

    # Remove cities from names
    name1_clean = remove_city_names(name1_clean)
    name2_clean = remove_city_names(name2_clean)

    # Normalize names for more precise comparison
    name1_clean = re.sub(r'\b(NAT|NATIONAL|SCHOOL|TECHNICAL|THEORETICAL|INDUSTRIAL)\b', '', name1_clean).strip()
    name2_clean = re.sub(r'\b(NAT|NATIONAL|SCHOOL|TECHNICAL|THEORETICAL|INDUSTRIAL)\b', '', name2_clean).strip()

    # Remove accidental duplicates
    name1_clean = re.sub(r'\s+', ' ', name1_clean).strip()
    name2_clean = re.sub(r'\s+', ' ', name2_clean).strip()

    # Check normal similarity
    if fuzz.ratio(name1_clean, name2_clean) > 85 or fuzz.partial_ratio(name1_clean, name2_clean) > 90:
        return True

    # Special handling for art schools
    if ("MUSIC" in name1_clean and "ART" in name1_clean) or ("MUSIC" in name2_clean and "ART" in name2_clean):
        if "ART" in name1_clean and "ART" in name2_clean:
            return True  # Considered equivalent if both contain "ART"

    # Special handling for Avram Iancu National College Câmpeni
    if "AVRAM IANCU" in name1_clean and "AVRAM IANCU" in name2_clean:
        return True

    # If one name is a substring of the other
    if name1_clean in name2_clean or name2_clean in name1_clean:
        return True

    return False


def read_excel_with_bold_marking(file_path, sheet_name):
    """ Read data from an Excel file and identify counties marked in bold. """
    wb = load_workbook(file_path, data_only=True)
    sheet = wb[sheet_name]

    data = []
    current_county = None

    for row in sheet.iter_rows():
        cell = row[0]
        if cell.font and cell.font.bold:
            current_county = normalize_text(cell.value)
        else:
            if current_county and cell.value:
                school = normalize_text(cell.value)
                num_students = row[1].value if len(row) > 1 and row[1].value else 0
                data.append((current_county, school, num_students))

    return pd.DataFrame(data, columns=["County", "School", "Number of Students"])


def group_similar_schools(df):
    """ Group similar schools only within the same county. """
    grouped_data = defaultdict(int)

    for county in df["County"].unique():
        df_county = df[df["County"] == county]

        for _, row in df_county.iterrows():
            school, num_students = row["School"], row["Number of Students"]
            found_match = False

            for existing_school in list(grouped_data.keys()):
                existing_county, existing_school_name = existing_school

                if existing_county != county:
                    continue

                if is_similar_school(existing_school_name, school):
                    grouped_data[existing_school] += num_students
                    found_match = True
                    break

            if not found_match:
                grouped_data[(county, school)] = num_students

    df_result = pd.DataFrame([(k[0], k[1], v) for k, v in grouped_data.items()],
                             columns=["County", "School", "Number of Students"])

    total_per_county = df_result.groupby("County")["Number of Students"].sum().to_dict()
    df_result["Total County"] = df_result["County"].map(total_per_county)

    return df_result


def save_to_excel(df, folder, filename):
    """ Save the data to an Excel file with counties grouped and the total per county in a single merged cell. """
    output_path = os.path.join(BASE_DIR, folder, "Result_" + filename)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="Result", index=False)
        ws = writer.sheets["Result"]

        current_county = None
        start_row = 2

        for row in range(2, len(df) + 2):
            if df.iloc[row - 2, 0] != current_county:
                if current_county is not None:
                    ws.merge_cells(start_row=start_row, start_column=1, end_row=row - 1, end_column=1)
                    ws.merge_cells(start_row=start_row, start_column=4, end_row=row - 1, end_column=4)
                current_county = df.iloc[row - 2, 0]
                start_row = row

    print(f"File successfully saved: {output_path}")


# ---- SCRIPT EXECUTION ----
if __name__ == "__main__":
    folder = input("Enter the folder: ").strip()
    filename = input("Enter the filename: ").strip()

    file_path = os.path.join(BASE_DIR, folder, filename)

    if os.path.exists(file_path):
        df = read_excel_with_bold_marking(file_path, sheet_name="Sheet2")
        df_grouped = group_similar_schools(df)
        save_to_excel(df_grouped, folder, filename)
    else:
        print("The file does not exist!")
